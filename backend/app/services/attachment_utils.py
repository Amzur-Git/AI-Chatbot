from __future__ import annotations

from pathlib import Path
from io import BytesIO
import re
from typing import Iterable

from fastapi import HTTPException, UploadFile

IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}
VIDEO_EXTENSIONS = {"mp4", "mov", "webm"}
TABLE_EXTENSIONS = {"csv", "xlsx"}
DOCUMENT_EXTENSIONS = {"pdf"}
FORMULA_EXTENSIONS = {"tex", "latex"}
CODE_EXTENSIONS = {"py", "js", "ts", "java", "cpp", "json", "sql"}

ALLOWED_EXTENSIONS = (
    IMAGE_EXTENSIONS
    | VIDEO_EXTENSIONS
    | TABLE_EXTENSIONS
    | DOCUMENT_EXTENSIONS
    | FORMULA_EXTENSIONS
    | CODE_EXTENSIONS
)


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]", "_", name)
    return cleaned[:180] or "upload.bin"


def infer_extension(upload_file: UploadFile) -> str:
    return Path(upload_file.filename or "").suffix.lower().lstrip(".")


def validate_upload(upload_file: UploadFile, size_bytes: int, max_size_mb: int, allowed_extensions: Iterable[str] | None = None) -> str:
    ext = infer_extension(upload_file)
    effective_allowed = set(allowed_extensions or ALLOWED_EXTENSIONS)

    if not ext or ext not in effective_allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '.{ext or 'unknown'}'. Allowed: {', '.join(sorted(effective_allowed))}",
        )

    limit = max_size_mb * 1024 * 1024
    if size_bytes > limit:
        raise HTTPException(status_code=413, detail=f"File too large. Max allowed is {max_size_mb}MB")

    return ext


def infer_category(ext: str) -> str:
    if ext in IMAGE_EXTENSIONS:
        return "image"
    if ext in VIDEO_EXTENSIONS:
        return "video"
    if ext in TABLE_EXTENSIONS:
        return "table"
    if ext in FORMULA_EXTENSIONS:
        return "formula"
    if ext in CODE_EXTENSIONS:
        return "code"
    return "document"


def extract_text_preview(content: bytes, ext: str, limit_chars: int = 2000) -> str | None:
    if ext not in (TABLE_EXTENSIONS | DOCUMENT_EXTENSIONS | FORMULA_EXTENSIONS | CODE_EXTENSIONS):
        return None

    if ext == "pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(BytesIO(content))
            parts: list[str] = []
            for page in reader.pages:
                page_text = (page.extract_text() or "").strip()
                if page_text:
                    parts.append(page_text)
                if sum(len(part) for part in parts) >= limit_chars:
                    break

            merged = "\n".join(parts).strip()
            if merged:
                return merged[:limit_chars]
        except Exception:
            pass

        # Fallback: try to recover visible ASCII/UTF-8-like strings from raw PDF bytes.
        try:
            raw_text = content.decode("latin-1", errors="ignore")
            cleaned = re.sub(r"[^\x20-\x7E\n\r\t]", " ", raw_text)
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            if cleaned:
                return cleaned[:limit_chars]
        except Exception:
            return None

        return None

    if ext == "xlsx":
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows: list[str] = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx > 30:
                    break
                values = [str(item).strip() for item in row[:20] if item is not None and str(item).strip()]
                if values:
                    rows.append(", ".join(values))
            workbook.close()
            preview = "\n".join(rows).strip()
            return preview[:limit_chars] if preview else None
        except Exception:
            return None

    try:
        text = content.decode("utf-8", errors="ignore")
    except Exception:
        return None

    preview = text.strip()
    if not preview:
        return None
    return preview[:limit_chars]
